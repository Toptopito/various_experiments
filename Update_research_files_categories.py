# This code updates the antidepressant research files from Dr. Alemi for ChatGPT completions
import openpyxl


def classify_entry(description):
    
    search_term = description.lower()

    # estrogen medications. 
    estrogen_medications = ['estrogen,con/m-progest acet', 'estrogen,ester/me-testosterone']

    for item in estrogen_medications:
        if item in search_term:
            classification = 'estrogen medication'
            return classification
      
    # hormonal contraception | estrogen
    if 'estrogen' in search_term:
        classification = 'hormonal contraception | estrogen medication'
        return classification
    
    # hormonal contraception 
    hormonal_contraceptions = ['desogestrel-ethinyl estradiol', 'ethinyl estradiol/drospirenone', \
                               'etonogestrel/ethinyl estradiol', 'levonorgestrel-ethin estradiol', \
                                   'norethindrone', 'norethindrone-ethin estradiol', \
                                       'norgestimate-ethinyl estradiol', 'estradiol']
        
    for item in hormonal_contraceptions:
        if item in search_term:
            classification = 'hormonal contraception'
            return classification

    # sedatives, tranquilizers, or muscle relaxants
    sedatives = ['eszopiclone', 'temazepam', 'zaleplon', 'zolpidem tartrate', 'alprazolam', 'diazepam', \
                 'lorazepam', 'baclofen', 'carisoprodol', 'cyclobenzaprine hcl', 'metaxalone', \
                     'tizanidine hcl', 'oxybutynin chloride', 'solifenacin succinate', 'gabapentin', \
                         'buspirone hcl', 'olanzapine', 'quetiapine fumarate', 'risperidone', 'ziprasidone hcl']

    for item in sedatives:
        if item in search_term:
            classification = 'sedative | tranquilizer | muscle relaxant'
            return classification

    # lipotropic medications 
    liprotic_medications = ['atorvastatin calcium', 'ezetimibe/simvastatin', 'fenofibrate', \
                            'fenofibrate nano crystallized', 'fenofibrate micronized', 'lovastatin', \
                                'pravastatin sodium', 'rosuvastatin calcium']

    for item in liprotic_medications:
        if item in search_term:
            classification = 'liprotic medication'
            return classification

    # antihypertensive or medications affecting the cardiovascular system
    antihypertensives = ['amlodipine', 'besylate/benazepril', 'benazepril hcl', 'hydralazine hcl', \
                         'lisinopril', 'lisinopril/hydrochlorothiazide', 'losartan/hydrochlorothiazide', \
                             'olmesartan/hydrochlorothiazide', 'ramipril', 'sildenafil citrate', \
                                 'valsartan/hydrochlorothiazide', 'amlodipine besylate', 'atenolol', \
                                     'metoprolol succinate', 'triamterene/hydrochlorothiazide', \
                                         'potassium chloride', 'gemfibrozil', 'simvastatin', 'verapamil hcl']

    for item in antihypertensives:
        if item in search_term:
            classification = 'antihypertensive'
            return classification

    # anticoagulants
    anticoagulants = ['apixaban', 'eliquis', 'dabigatran', 'pradaxa', 'edoxaban', 'lixiana', 'rivaroxaban', \
                      'xarelto', 'warfarin', 'coumadin']

    for item in anticoagulants:
        if item in search_term:
            classification = 'anticoagulant'
            return classification

    # antihistamines
    antihistamines = ['cetirizine hcl', 'desloratadine', 'fexofenadine hcl', 'fexofenadine/pseudoephedrine', \
                      'hydroxyzine hcl', 'hydroxyzine pamoate', 'loratadine', 'promethazine hcl', ]



    for item in antihistamines:
        if item in search_term:
            classification = 'antihistamine'
            return classification

    # antibiotics
    antibiotics = ['amoxicillin', 'clindamycin hcl', 'clindamycin phosphate', 'penicillin v potassium', \
                   'metronidazole', 'cefprozil', 'clarithromycin', 'sulfamethoxazole/trimethoprim', \
                       'doxycycline hyclate', 'mupirocin']

    for item in antibiotics:
        if item in search_term:
            classification = 'antibiotic'
            return classification

    # medications for nausea
    nausea_medications = ['meclizine hcl', 'metoclopramide hcl', 'prochlorperazine maleate', 'ondansetron']

    for item in nausea_medications:
        if item in search_term:
            classification = 'nausea medication'
            return classification

    # medications for ulcer
    ulcer_medications = ['someprazole magnesium', 'famotidine', 'lansoprazole', 'omeprazole', 'pantoprazole sodium', \
                         'rabeprazole sodium']

    for item in ulcer_medications:
        if item in search_term:
            classification = 'ulcer medication'
            return classification

    # analgesic medications 
    analgesic_medications = ['acetaminophen with codeine', 'fentanyl', 'hydrocodone/acetaminophen', \
                             'hydrocodone/ibuprofen', 'morphine sulfate', 'oxycodone hcl', \
                                 'oxycodone hcl/acetaminophen', 'propoxyphene nap/acetaminophen', \
                                     'tramadol hcl', 'tramadol hcl/acetaminophen', 'rizatriptan benzoate', \
                                         'sumatriptan succinate', 'divalproex sodium']

    for item in analgesic_medications:
        if item in search_term:
            classification = 'analgesic medication'
            return classification

    # cold, cough, or nasal medicines
    cold_medications = ['hydrocodone/chlorphen p-stirex', 'phenylephrine/hydrocodone/cpm', \
                        'codeine phosphate/guaifenesin', 'guaifenesin/p-ephed hcl', \
                            'guaifenesin/phenylephrine hcl', 'ciprofloxacin hcl', 'epinephrine', \
                                'triamcinolone acetonide']

    for item in cold_medications:
        if item in search_term:
            classification = 'cold | cough | nasal'
            return classification

    # bronchial dilation medications 
    bronchial_dilation_medications = ['albuterol sulfate', 'fluticasone propion/salmeterol', 'montelukast sodium']

    for item in bronchial_dilation_medications:
        if item in search_term:
            classification = 'bronchial dilation medication'
            return classification

    # medications for Parkinson disease. 
    Parkinson_medications = ['carbidopa/levodopa', 'pramipexole di-hcl', 'ropinirole hcl']

    for item in Parkinson_medications:
        if item in search_term:
            classification = 'Parkinson disease medication'
            return classification

    # Central Nervous System stimulant medications 
    CNS_stimulants = ['methylphenidate hcl', 'dexmethylphenidate dextroamphetamine-amphetamine salts', \
                                 'lisdexamfetamine', 'varenicline tartrate']

    for item in CNS_stimulants:
        if item in search_term:
            classification = 'CNS stimulant'
            return classification

    # antidiabetics 
    antidiabetics = ['insulin lispro', 'metformin hcl', 'needles, insulin disposable']

    for item in antidiabetics:
        if item in search_term:
            classification = 'antidiabetic'
            return classification

    # eye medications
    eye_medications = ['azithromycin', 'diclofenac sodium', 'latanoprost']

    for item in eye_medications:
        if item in search_term:
            classification = 'eye medication'
            return classification

    # antifungal medications 
    antifungals = ['conjugated fluconazole', 'clotrimazole/betamethasone dip', 'ketoconazole', \
                   'terconazole']

    for item in antifungals:
        if item in search_term:
            classification = 'antifungal'
            return classification

    # immunosuppresive drug | chemotherapy
    if 'methotrexate sodium' in search_term:
        classification = 'immunosuppresive drug | chemotherapy'
        return classification

    # immunosuppresive drug
    if 'hydroxychloroquine sulfate' in search_term:
        classification = 'immunosuppresive drug'
        return classification    

    # chemotherapy | hormonal therapy
    if 'tamoxifen citrate' in search_term:
        classification = 'chemotherapy | hormonal therapy'
        return classification   
       
    # anti-steroidal
    if 'hydrocortisone' in search_term:
        classification = 'anti-steroidal'
        return classification   

    # steriod
    if 'methylprednisolone' in search_term:
        classification = 'steriod'
        return classification   

    # anesthetic
    if 'lidocaine' in search_term:
        classification = 'anesthetic'
        return classification   

    # dementia medication
    if 'donepezil hcl' in search_term:
        classification = 'dementia medication'
        return classification   

    # laxative
    laxatives = ['lactulose', 'polyethylene glycol 3350']

    for item in laxatives:
        if item in search_term:
            classification = 'laxative'
            return classification 

    # antivirals
    antivirals = ['oseltamivir phosphate', 'valacyclovir hcl']

    for item in antivirals:
        if item in search_term:
            classification = 'antiviral'
            return classification 

    # thyroid medication
    if 'levothyroxine sodium' in search_term:
        classification = 'thyroid medication'
        return classification   

    # osteoporosis medication
    if 'risedronate sodium' in search_term:
        classification = 'osteoporosis medication'
        return classification   
        
    # antibacterial
    if 'silver sulfadiazine' in search_term:
        classification = 'antibacterial'
        return classification  

    # urinary retention
    if 'tamsulosin' in search_term:
        classification = 'urinary retention'
        return classification  
    
    # antiarthritics
    antiarthritics = ['allopurinol', 'celecoxib', 'ibuprofen', 'meloxicam', 'nabumetone', 'naproxen', 'rofecoxib']

    for item in antiarthritics:
        if item in search_term:
            classification = 'antiarthritic'
            return classification 

    # dermatological
    if 'metronidazole' in search_term:
        classification = 'dermatological'
        return classification  
    
    # amphetamine preparations
    if 'dextroamphetamine/amphetamine' in search_term:
        classification = 'amphetamine preparation'
        return classification  

    # anticonvulsants
    anticonvulsants = ['clonazepam', 'lamotrigine', 'pregabalin', 'topiramate']

    for item in anticonvulsants:
        if item in search_term:
            classification = 'anticonvulsant'
            return classification      

    # diuretic
    if 'triamterene/hydrochlorothiazid' in search_term:
        classification = 'diuretic'
        return classification 
    
    # miscellaneous
    miscellaneous = ['fluconazole', 'testosterone']

    for item in miscellaneous:
        if item in search_term:
            classification = 'miscellaneous'
            return classification 
        
    # psychostimulant-antidepressant
    psychostimulants = ['amitriptyline hcl', 'aripiprazole', 'bupropion', 'bupropion hcl', \
                        'citalopram hydrobromide', 'clonidine hcl', 'desvenlafaxine succinate', \
                            'doxepin hcl', 'duloxetine hcl', 'escitalopram oxalate', 'fluoxetine hcl', \
                                'fluvoxamine maleate', 'imipramine hcl', 'lithium carbonate', 'mirtazapine', \
                                    'modafinil', 'nefazodone hcl', 'nortriptyline hcl', 'paroxetine hcl', \
                                        'sertraline hcl', 'trazodone hcl', 'venlafaxine hcl', 'vilazodone hcl']

    for item in psychostimulants:
        if item in search_term:
            classification = 'psychostimulant-antidepressant'
            return classification         

    return 'None'


def fill_category_column():
    # define constants
    revised_coeff_file_path = 'C:/Users/vladc/OneDrive/Documents/GMU Research/AI Chat for Depression/Reference Files/Revised Rem Varables Descriptions.xlsx'
    revised_coeff_sheet_name = 'Revised'
    description_column = 'J'
    category_column = 16
    revised_coeff_start_row = 2
    
    revised_coeff_file = openpyxl.load_workbook(revised_coeff_file_path)
    revised_coeff_sheet = revised_coeff_file[revised_coeff_sheet_name]
    
    # loop through coeff file and fill in the category column
    for row in range(revised_coeff_start_row, revised_coeff_sheet.max_row+1):
        description_cell = "{}{}".format(description_column, row)
        description = revised_coeff_sheet[description_cell].value
   
        category = classify_entry(description)
        
        category_cell = revised_coeff_sheet.cell(row, category_column)
        category_cell.value = category
        
    revised_coeff_file.save(revised_coeff_file_path)

    return

def main():
    
    fill_category_column()
    
    return

if __name__ == "__main__":
    main()