# This code updates the antidepressant research files from Dr. Alemi for ChatGPT completions
import openpyxl


def fill_pivot_completions(pivot_sheet):
    # constants
    pivot_completion_column = 21
    pivot_start_row = 6
    pivot_success_rates_columns = {'C' : 'AMITRIPTYLINE', 'D' : 'BUPROPION', 'E' : 'CITALOPRAM', 'F' : 'DESVENLAFAXINE', \
                                   'G' : 'DOXEPIN', 'H' : 'DULOXETINE', 'I' : 'ESCITALOPRAM', 'J' : 'FLUOXETINE', \
                                       'K' : 'MIRTAZAPINE', 'L' : 'NORTRIPTYLINE', 'M' : 'OTHER', 'N' : 'PAROXETINE', \
                                           'O' : 'ROPINIROLE', 'P' : 'SERTRALINE', 'Q' : 'TRAZODONE', 'R' : 'VENLAFAXINE'}
    min_success_rate = 0.1
    max_diff = 0.05

    # loop through workbook pivot sheet
    for row in range(pivot_start_row, pivot_sheet.max_row+1):
    #for row in range(6, 11): #test rows only    
    
        # list values of success rates for each category
        success_values = dict()
        for col in pivot_success_rates_columns:
            # all success rates in a row
            success_rate_cell = "{}{}".format(col, row)
            success_rate = pivot_sheet[success_rate_cell].value
            success_values[pivot_success_rates_columns[col]] = success_rate
        
        # get top 1 and top 2 success rates
        top1_val = 0
        top1_name = 'None'
        top2_val = 0
        top2_name = 'None'
        for item in success_values:
            if success_values[item] is None:
                continue
                        
            if success_values[item] > top1_val:
                top1_val = success_values[item]
                top1_name = item.title()

            if success_values[item] > top2_val and success_values[item] < top1_val:
                top2_val = success_values[item]
                top2_name = item.title()       
                  
        # create recommendation
        if top1_val > min_success_rate and top2_val > min_success_rate:
            diff = top1_val - top2_val
            if abs(diff) < max_diff:
                recommendation = top1_name + ' | ' + top2_name
            else:
                recommendation = top1_name
        elif top1_val > min_success_rate:
            recommendation = top1_name
        else:
            recommendation = 'None'
        
        # write recommendation in completion column
        completion_cell = pivot_sheet.cell(row, pivot_completion_column)
        completion_cell.value = recommendation
        
    return


def create_completions_dict(pivot_sheet):
    # constants
    pivot_start_row = 6
    pivot_strata_sort_column = 'B'
    pivot_completion_column = 'U'
    
    completions_dict = dict()
    # loop through workbook pivot sheet
    for row in range(pivot_start_row, pivot_sheet.max_row+1):
    #for row in range(6, 11): #test rows only    
    
        # get strata values
        strata_cell = "{}{}".format(pivot_strata_sort_column, row)
        strata = pivot_sheet[strata_cell].value
        
        # get completion value
        completion_cell = "{}{}".format(pivot_completion_column, row)
        completion = pivot_sheet[completion_cell].value
        
        #create dict entry
        completions_dict[strata] = completion
    
    return completions_dict


def fill_flat_completions(flat_sheet, completions_dict):
    # constants
    flat_completion_column = 11
    flat_start_row = 2
    flat_strata_sort_column = 'H'
    
    # loop through workbook flat sheet
    for row in range(flat_start_row, flat_sheet.max_row+1):
    #for row in range(6, 11): #test rows only
        
        # get strata value
        strata_cell = "{}{}".format(flat_strata_sort_column, row)
        strata = flat_sheet[strata_cell].value
        
        # get completion value
        completion = completions_dict[strata]
        
        # write completion
        completion_cell = flat_sheet.cell(row, flat_completion_column)
        completion_cell.value = completion        
    
    return


def fill_completions(pivot_sheet, flat_sheet):
       
    fill_pivot_completions(pivot_sheet)
    
    completions_dict = create_completions_dict(pivot_sheet)
    
    fill_flat_completions(flat_sheet, completions_dict)
    
    return


def fill_pivot_prompts(pivot_sheet):
    # constants
    pivot_prompt_column = 20
    pivot_unmatched_column = 22
    pivot_start_row = 6
    pivot_strata_column = 'B'

    # loop through workbook pivot sheet
    for row in range(pivot_start_row, pivot_sheet.max_row+1):
    #for row in range(11, 111): #test rows only    
    
        # read strata
        strata_cell = "{}{}".format(pivot_strata_column, row)
        strata = pivot_sheet[strata_cell].value
        
        prompt_segments = [] # initialize the prompt_segments
        unmatched_segments = [] # list any remaining gaps in parsing
        procedure_segments = [] # for ICD9:V and CPT4 make sure all procedures are recorded together
        symptom_segments = [] # for ICD9 make sure all symptoms are recorded together       
        
        if strata is None:
            continue
        
        # parse strata with attributes split using '|'        
        strata_segments = strata.split('|')
        
        # loop into strata segments
        for segment in strata_segments:
            # age
            if 'age:' in segment.lower():
                age_range = segment.split(':')[1].strip()
                age_low = age_range.split('-')[0].strip()
                age_high = age_range.split('-')[1].strip()
                prompt_segments.append(f"{age_low} to {age_high} years old")
            
            # gender
            elif 'gender:' in segment.lower():
                gender = segment.split(':')[1].strip().lower()
                prompt_segments.append(f"{gender}")
            
            # procedures
            elif 'icd9:v' in segment.lower() or 'cpt4' in segment.lower():
                procedure = segment.split('(')[0].strip().lower()
                
                # remove other texts
                if ',' in procedure:
                    procedure = procedure.split(',')[0].strip()
                    
                procedure_segments.append(f"{procedure}")
                
            # symptoms
            elif 'icd9' in segment.lower():
                symptom = segment.split('(')[0].strip().lower()
                
                mdd_re_flag = False #check for recurrent mdd
                if 'major depressive disorder, recurrent episode' in symptom:
                    mdd_re_flag = True
                    
                # remove other texts
                if ',' in symptom:
                    symptom = symptom.split(',')[0].strip()
                
                if 'of unspecified site' in symptom:
                    symptom = symptom.split('of unspecified site')[0].strip()

                if mdd_re_flag:
                    symptom_segments.append("recurrent episodes of major depressive disorder")
                else:
                    symptom_segments.append(f"{symptom}")
                    
            # antidepressants
            
            # last antidepressant
            elif 'last ad:' in segment.lower():
                if 'same & no remiss' in segment.lower():
                    prompt_segments.append("last antidepressant is the same but with no remission")
                elif 'same & remiss' in segment.lower():
                    prompt_segments.append("last antidepressant is the same with remission")
                elif 'diff & no remiss' in segment.lower():
                    prompt_segments.append("last antidepressant is different but with no remission")
                elif 'diff & remiss' in segment.lower():
                    prompt_segments.append("last antidepressant is different with remission")
            
            # unmatched
            else:
                unmatched_segments.append(f"<{segment}>")
                
        
        prompt = ' '.join(prompt_segments)
        unmatched = ' '.join(unmatched_segments)
        
        symptoms = ""
        for i in range(len(symptom_segments)):
            if i == 0:
                symptoms += f"with {symptom_segments[i]}"
            else:
                symptoms += f" and {symptom_segments[i]}"
        
        procedures = ""
        for i in range(len(procedure_segments)):
            if i == 0:
                procedures += f"underwent {procedure_segments[i]}"
            else:
                procedures += f" and {procedure_segments[i]}"
        
        # write the prompt and unmatched in their columns
        prompt_cell = pivot_sheet.cell(row, pivot_prompt_column)
        prompt_cell.value = prompt + ' ' + symptoms + ' ' + procedures
        unmatched_cell = pivot_sheet.cell(row, pivot_unmatched_column)
        unmatched_cell.value = unmatched   
       
    return


def create_prompts_dict(pivot_sheet):
    # constants
    pivot_start_row = 6
    pivot_strata_sort_column = 'B'
    pivot_prompt_column = 'T'
    
    prompts_dict = dict()
    # loop through workbook pivot sheet
    for row in range(pivot_start_row, pivot_sheet.max_row+1):
    #for row in range(6, 11): #test rows only    
    
        # get strata values
        strata_cell = "{}{}".format(pivot_strata_sort_column, row)
        strata = pivot_sheet[strata_cell].value
        
        # get completion value
        prompt_cell = "{}{}".format(pivot_prompt_column, row)
        prompt = pivot_sheet[prompt_cell].value
        
        #create dict entry
        prompts_dict[strata] = prompt
    
    return prompts_dict


def fill_flat_prompts (flat_sheet, prompts_dict):
    # constants
    flat_prompt_column = 10
    flat_start_row = 2
    flat_strata_sort_column = 'H'
    
    # loop through workbook flat sheet
    for row in range(flat_start_row, flat_sheet.max_row+1):
    #for row in range(6, 11): #test rows only
        
        # get strata value
        strata_cell = "{}{}".format(flat_strata_sort_column, row)
        strata = flat_sheet[strata_cell].value
        
        # get completion value
        prompt = prompts_dict[strata]
        
        # write completion
        prompt_cell = flat_sheet.cell(row, flat_prompt_column)
        prompt_cell.value = prompt 
        
    return


def fill_prompts(pivot_sheet, flat_sheet):
    
    fill_pivot_prompts(pivot_sheet)
    
    prompts_dict = create_prompts_dict(pivot_sheet)
    
    fill_flat_prompts(flat_sheet, prompts_dict)
    
    return


def main():

    # define constants
    workbook_file_path = 'C:/Users/vladc/OneDrive/Documents/GMU Research/AI Chat for Depression/Reference Files/Subgroups and Optimal Antidepressants.xlsx'
    pivot_sheet_name = 'Prompts and Completions'
    flat_sheet_name = 'Prompts Completions Flat'
    
    workbook = openpyxl.load_workbook(workbook_file_path)
    pivot_sheet = workbook[pivot_sheet_name]
    flat_sheet = workbook[flat_sheet_name]
    
    fill_completions(pivot_sheet, flat_sheet)
    
    fill_prompts(pivot_sheet, flat_sheet)

    workbook.save(workbook_file_path)    
    
    return

if __name__ == "__main__":
    main()